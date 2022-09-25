# Dependencies
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# Creating a Funtion, Loading Workbook


def process_workbook(filename):
    wb = xl.load_workbook(filename)

    # Creating Values
    sheet = wb["Sheet1"]

    # For loop correcting Prices
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        correctted_price_cell = sheet.cell(row, 4)
        correctted_price_cell.value = corrected_price

    # Creating Value reference
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    # Creating BarChart on Workbook
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "e2")

    # Saving New Workbook
    wb.save(filename)
